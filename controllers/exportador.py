"""
Módulo de Exportação Contábil (Controller).

Este módulo contém as regras de formatação e expurgo necessárias para
gerar o arquivo CSV customizado (que leva extensão .txt) de Ofertas,
permitindo a importação manual em lote no sistema legado do SIGA.
"""

import logging

class Exportador:
    """
    Responsável por processar e exportar transações bancárias para formato texto.
    
    Ele filtra as transações não pertinentes a Ofertas de igrejas (ex: Saques,
    Rendimentos, etc.) e formata os números no padrão brasileiro exigido pelo
    importador ASP do SIGA.
    """
    
    @staticmethod
    def gerar_txt_ofertas(transacoes, filepath):
        """
        Compilador léxico em conformidade com as regras IT.TES.05.
        
        Realiza a leitura da lista nativa de transações OFX carregada, itera
        sobre as descrições e expurga registros indesejados usando uma 
        cartilha de Blacklist.
        
        Args:
            transacoes (list): Lista de dicionários contendo transações do OFX.
            filepath (str): Caminho físico absoluto onde o arquivo TXT será salvo.
            
        Returns:
            tuple: Uma tupla contendo 3 elementos:
                   - sucesso (bool): True se gerou arquivo, False se não sobrou dados.
                   - linhas_validas (int): Quantidade de ofertas mantidas no arquivo.
                   - qtd_descartadas (int): Quantidade de transações ignoradas.
            
        Raises:
            Exception: Caso ocorra problema de I/O na gravação do disco ou falta
                       de permissão na pasta de destino escolhida.
        """
        # Proteção contra lista vazia ou nula
        if not transacoes:
            return False, 0, 0

        # Palavras bloqueadas que categorizam lançamentos que NÃO são Ofertas puras.
        # Por exemplo: Rendimentos de aplicação, saques aplicados e dinheiro em 
        # caixa já lançado direto (depositos ou congregação).
        palavras_bloqueadas = ["APLICA", "RESGATE", "RENDIMENT", "DEPOSITO", "RESG.", "DEP ", "CONGREGACAO", "TRANSFERENCIA"]

        linhas_limpas = []
        qtd_descartadas = 0

        for tx in transacoes:
            valor = tx.get("valor", 0.0)
            descricao = tx.get("descricao", "").upper()
            data = tx.get("data", "")

            # COND. 1: Somente valores POSITIVOS entram no lançamento puro (sem "-").
            # Valores negativos (débitos) são ignorados pois o SIGA não aceita estornos neste arquivo.
            if valor <= 0:
                qtd_descartadas += 1
                continue
            
            # COND. 2: Nenhuma palavra bloqueada pode fazer parte da descrição da string OFX.
            if any(palavra in descricao for palavra in palavras_bloqueadas):
                qtd_descartadas += 1
                continue

            # FORMATAR VALOR (sem separador de milhar, virgula p/ centavos. ex: 1530,50)
            # O SIGA legado em ASP classic quebra se mandarmos '1.530,50' ou '1530.50'
            valor_fmt = f"{valor:.2f}".replace(".", ",")
            
            # Formato SIGA Csv: DATA;HISTÓRICO;VALOR
            linha = f"{data};{tx.get('descricao', '').strip()};{valor_fmt}"
            linhas_limpas.append(linha)

        if not linhas_limpas:
            return False, 0, qtd_descartadas

        try:
            # Utiliza utf-8 padrão (sem BOM) para evitar que o importador web ASP do SIGA
            # leia a assinatura invisível do arquivo e devolva um erro bizarro na linha 1
            # com os caracteres de controle 'ï»¿'.
            with open(filepath, 'w', encoding='utf-8') as f:
                for l in linhas_limpas:
                    f.write(l + "\n")
            return True, len(linhas_limpas), qtd_descartadas
        except Exception as e:
            logging.error(f"Erro ao salvar arquivo TXT das ofertas: {e}", exc_info=True)
            raise e

    @staticmethod
    def gerar_excel_lote(telemetria, dados_lote, filepath):
        """
        Gera uma planilha Excel formatada (.xlsx) com o resumo do lote de conciliação,
        incluindo transações processadas, status de conferência e pendências.
        """
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        
        wb = openpyxl.Workbook()
        
        # 1. Planilha de Resumo Geral
        ws_resumo = wb.active
        ws_resumo.title = "Resumo do Lote"
        ws_resumo.views.sheetView[0].showGridLines = True
        
        # Cores e Estilos (Alinhado com a identidade visual)
        cor_header = PatternFill(start_color="438EB9", end_color="438EB9", fill_type="solid")
        cor_zebra = PatternFill(start_color="F9FBFD", end_color="F9FBFD", fill_type="solid")
        fonte_titulo = Font(name="Segoe UI", size=16, bold=True, color="438EB9")
        fonte_header = Font(name="Segoe UI", size=11, bold=True, color="FFFFFF")
        fonte_normal = Font(name="Segoe UI", size=10)
        fonte_bold = Font(name="Segoe UI", size=10, bold=True)
        
        # Título
        ws_resumo["A1"] = "AutoSIGA - Relatório de Produtividade do Lote"
        ws_resumo["A1"].font = fonte_titulo
        ws_resumo.row_dimensions[1].height = 30
        
        # Métricas Gerais
        ws_resumo["A3"] = "Métrica"
        ws_resumo["B3"] = "Valor"
        for col in ["A", "B"]:
            ws_resumo[f"{col}3"].fill = cor_header
            ws_resumo[f"{col}3"].font = fonte_header
            ws_resumo[f"{col}3"].alignment = Alignment(horizontal="center")
            
        dados_metrica = [
            ("Total de Extratos Processados", telemetria.get("total_contas", 0)),
            ("Transações no Extrato (OFX)", telemetria.get("ofx_itens", 0)),
            ("Transações Localizadas no SIGA", telemetria.get("siga_itens", 0)),
            ("Lançamentos Injetados", telemetria.get("injecoes", 0)),
            ("Pendências Finais não Conciliadas", telemetria.get("pendentes", 0)),
            ("Volume Financeiro Conciliado (R$)", telemetria.get("volume_financeiro", 0.0)),
        ]
        
        for idx, (m, v) in enumerate(dados_metrica, start=4):
            ws_resumo[f"A{idx}"] = m
            ws_resumo[f"B{idx}"] = v
            ws_resumo[f"A{idx}"].font = fonte_normal
            ws_resumo[f"B{idx}"].font = fonte_bold
            if "Volume" in m:
                ws_resumo[f"B{idx}"].number_format = 'R$ #,##0.00'
                ws_resumo[f"B{idx}"].alignment = Alignment(horizontal="right")
            else:
                ws_resumo[f"B{idx}"].alignment = Alignment(horizontal="right")
            
            # Linha zebrada
            if idx % 2 == 0:
                ws_resumo[f"A{idx}"].fill = cor_zebra
                ws_resumo[f"B{idx}"].fill = cor_zebra
                
        # 2. Planilha de Detalhamento por Extrato
        ws_detalhe = wb.create_sheet(title="Detalhamento")
        ws_detalhe.views.sheetView[0].showGridLines = True
        
        headers_det = ["Localidade", "Tipo Extrato", "Conta Bancária", "Total Transações", "Transações SIGA", "Injetados", "Status Final"]
        ws_detalhe.row_dimensions[1].height = 25
        for col_idx, text in enumerate(headers_det, start=1):
            cell = ws_detalhe.cell(row=1, column=col_idx, value=text)
            cell.fill = cor_header
            cell.font = fonte_header
            cell.alignment = Alignment(horizontal="center", vertical="center")
            
        for row_idx, d in enumerate(dados_lote, start=2):
            ws_detalhe.row_dimensions[row_idx].height = 20
            loc = f"{d.get('tipo_adm', '')} - {d.get('nome_adm', '')}"
            tipo = d.get('tipo_extrato', 'CONTA CORRENTE')
            conta = d.get('conta_id', '')
            total_tx = len(d.get('transacoes', []))
            total_siga = len(d.get('extrato_siga', []))
            
            # Tenta pegar a quantidade injetada de forma condicional se existirem pendências
            inj = total_tx - len(d.get('pendentes', [])) if 'pendentes' in d else total_tx
            status = "Conciliado 100%" if not d.get('pendentes') else f"{len(d.get('pendentes'))} pendentes"
            
            ws_detalhe.cell(row=row_idx, column=1, value=loc).font = fonte_normal
            ws_detalhe.cell(row=row_idx, column=2, value=tipo).font = fonte_normal
            ws_detalhe.cell(row=row_idx, column=3, value=conta).font = fonte_normal
            ws_detalhe.cell(row=row_idx, column=4, value=total_tx).font = fonte_normal
            ws_detalhe.cell(row=row_idx, column=5, value=total_siga).font = fonte_normal
            ws_detalhe.cell(row=row_idx, column=6, value=inj).font = fonte_normal
            
            status_cell = ws_detalhe.cell(row=row_idx, column=7, value=status)
            status_cell.font = fonte_bold
            if "100%" in status:
                status_cell.fill = PatternFill(start_color="DFF0D8", end_color="DFF0D8", fill_type="solid")
                status_cell.font = Font(name="Segoe UI", size=10, bold=True, color="3C763D")
            else:
                status_cell.fill = PatternFill(start_color="F2DEDE", end_color="F2DEDE", fill_type="solid")
                status_cell.font = Font(name="Segoe UI", size=10, bold=True, color="A94442")
                
            if row_idx % 2 == 0:
                for c in range(1, 7):
                    ws_detalhe.cell(row=row_idx, column=c).fill = cor_zebra
                    
        # Redimensionamento dinâmico das colunas
        for sheet in [ws_resumo, ws_detalhe]:
            for col in sheet.columns:
                max_len = 0
                for cell in col:
                    if cell.value:
                        max_len = max(max_len, len(str(cell.value)))
                col_letter = openpyxl.utils.get_column_letter(col[0].column)
                sheet.column_dimensions[col_letter].width = max(max_len + 3, 12)
                
        wb.save(filepath)

